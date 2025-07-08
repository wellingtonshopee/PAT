# patrimonio/views.py
import csv
from django.http import HttpResponse, JsonResponse # Adicionado JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse # Certifique-se que reverse está aqui
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.db.models import Q # Para consultas OR
from django.core.paginator import Paginator
from django.utils import timezone # Para timezone.now() e timedelta
from datetime import timedelta # Importe timedelta para o filtro de data
from .models import ItemPatrimonio, CategoriaPatrimonio, LocalizacaoPatrimonio, MovimentacaoPatrimonio
from .filters import ItemPatrimonioFilter
import base64 
from django.views.decorators.http import require_POST

import qrcode
from io import BytesIO


from .models import ItemPatrimonio, CategoriaPatrimonio, LocalizacaoPatrimonio, MovimentacaoPatrimonio
from .forms import ItemPatrimonioForm, PatrimonioFilterForm, BaixaPatrimonioForm, TransferenciaPatrimonioForm, MovimentacaoFilterForm

# Importar openpyxl para exportação de Excel
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

User = get_user_model()
# A linha 'ser = get_user_model()' não é necessária, pois 'User' já faz o mesmo. Removida para clareza.


@login_required
def gerar_qrcode_patrimonio(request, item_id):
    item = get_object_or_404(ItemPatrimonio, pk=item_id)
    
    # URL completa que será codificada no QR Code
    # Exemplo: http://127.0.0.1:8000/patrimonio/inventario/coletar/?codigo_patrimonial=SEUCODIGO
    url_para_qrcode = request.build_absolute_uri(
        reverse('patrimonio:coleta_inventario') + f'?codigo_patrimonial={item.codigo_patrimonial}'
    )

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url_para_qrcode)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = BytesIO()
    img.save(buffer, format="PNG") # Salva a imagem PNG no buffer
    
    return HttpResponse(buffer.getvalue(), content_type="image/png")


@login_required
def coleta_inventario(request):
    item_encontrado = None
    # Pega o código patrimonial da requisição GET (se existir)
    codigo_patrimonial_preenchido = request.GET.get('codigo_patrimonial', '').strip() 
    
    if codigo_patrimonial_preenchido:
        try:
            # Tenta encontrar o item pelo código patrimonial
            item_encontrado = ItemPatrimonio.objects.get(codigo_patrimonial__iexact=codigo_patrimonial_preenchido)
            messages.success(request, f"Item '{item_encontrado.nome}' ({item_encontrado.codigo_patrimonial}) encontrado.")
        except ItemPatrimonio.DoesNotExist:
            messages.error(request, f"Nenhum item encontrado com o código: {codigo_patrimonial_preenchido}.")
            item_encontrado = None # Garante que a variável seja None se não encontrar
        except Exception as e:
            # Captura outros erros inesperados durante a busca
            messages.error(request, f"Ocorreu um erro ao buscar o item: {e}")
            item_encontrado = None

    # Obter todas as localizações para o campo 'Nova Localização'
    localizacoes = LocalizacaoPatrimonio.objects.all().order_by('nome')
    
    # Obter usuários disponíveis para o campo 'Novo Responsável'
    # Você pode filtrar esses usuários se necessário (ex: apenas funcionários)
    users_disponiveis = User.objects.filter(is_active=True).order_by('first_name', 'last_name')

    context = {
        'codigo_patrimonial_preenchido': codigo_patrimonial_preenchido, # Para manter o campo preenchido após a busca
        'item_encontrado': item_encontrado,
        'localizacoes': localizacoes,
        'users_disponiveis': users_disponiveis,
    }
    return render(request, 'patrimonio/coleta_inventario.html', context)


# Exemplo de como seria a view confirmar_inventario (se você já não a tem)
# Ela é chamada quando o formulário de 'Registrar Presença / Atualizar Dados' é submetido
@login_required
@require_POST # Garante que esta view só aceita requisições POST
def confirmar_inventario(request):
    item_id = request.POST.get('item_id')
    nova_localizacao_id = request.POST.get('nova_localizacao')
    novo_responsavel_id = request.POST.get('novo_responsavel')
    observacoes_inventario = request.POST.get('observacoes_inventario')

    try:
        item = get_object_or_404(ItemPatrimonio, pk=item_id)
        
        # Registrar a movimentação antes de atualizar o item, se houver mudança
        if nova_localizacao_id and int(nova_localizacao_id) != item.localizacao.id:
            nova_localizacao_obj = get_object_or_404(LocalizacaoPatrimonio, pk=nova_localizacao_id)
            MovimentacaoPatrimonio.objects.create(
                item=item,
                tipo_movimentacao='TRANSFERENCIA', # Ou 'INVENTARIO_MOV' se for um tipo específico
                data_movimentacao=timezone.now(),
                localizacao_origem=item.localizacao,
                localizacao_destino=nova_localizacao_obj,
                observacoes=f"Transferência durante coleta de inventário. Nova localização: {nova_localizacao_obj.nome}. {observacoes_inventario}",
                realizada_por=request.user
            )
            item.localizacao = nova_localizacao_obj
            messages.info(request, f"Localização de '{item.nome}' atualizada para {nova_localizacao_obj.nome}.")

        if novo_responsavel_id and int(novo_responsavel_id) != (item.responsavel_atual.id if item.responsavel_atual else None):
            novo_responsavel_obj = get_object_or_404(User, pk=novo_responsavel_id)
            # Você pode registrar uma movimentação específica para mudança de responsável, se tiver um tipo
            # ou incluir nas observações da movimentação de localização se for a única.
            item.responsavel_atual = novo_responsavel_obj
            messages.info(request, f"Responsável de '{item.nome}' atualizado para {novo_responsavel_obj.get_full_name()}.")
        
        # Atualizar a data da última atualização para registrar a presença no inventário
        item.data_ultima_atualizacao = timezone.now()
        item.observacoes = observacoes_inventario # Ou adicionar/concatenar a uma observação de inventário
        item.save()

        messages.success(request, f"Inventário do item '{item.nome}' confirmado e dados atualizados com sucesso!")
        
    except ItemPatrimonio.DoesNotExist:
        messages.error(request, "Item não encontrado para atualização.")
    except LocalizacaoPatrimonio.DoesNotExist:
        messages.error(request, "Nova localização inválida.")
    except User.DoesNotExist:
        messages.error(request, "Novo responsável inválido.")
    except Exception as e:
        messages.error(request, f"Ocorreu um erro inesperado ao confirmar o inventário: {e}")

    return redirect('patrimonio:coleta_inventario') # Redireciona de volta para a página de coleta


@login_required
def listar_itens_patrimonio(request):
    itens_patrimonio = ItemPatrimonio.objects.all()
    filter_form = PatrimonioFilterForm(request.GET)

    if filter_form.is_valid():
        search_query = filter_form.cleaned_data.get('search_query')
        categoria = filter_form.cleaned_data.get('categoria')
        localizacao = filter_form.cleaned_data.get('localizacao')
        estado_conservacao = filter_form.cleaned_data.get('estado_conservacao')
        status = filter_form.cleaned_data.get('status')
        data_aquisicao_inicio = filter_form.cleaned_data.get('data_aquisicao_inicio')
        data_aquisicao_fim = filter_form.cleaned_data.get('data_aquisicao_fim')

        if search_query:
            itens_patrimonio = itens_patrimonio.filter(
                Q(nome__icontains=search_query) |
                Q(codigo_patrimonial__icontains=search_query) |
                Q(numero_serie__icontains=search_query) |
                Q(descricao__icontains=search_query)
            )
        if categoria:
            itens_patrimonio = itens_patrimonio.filter(categoria=categoria)
        if localizacao:
            itens_patrimonio = itens_patrimonio.filter(localizacao=localizacao)
        if estado_conservacao:
            itens_patrimonio = itens_patrimonio.filter(estado_conservacao=estado_conservacao)
        if status:
            itens_patrimonio = itens_patrimonio.filter(status=status)
        if data_aquisicao_inicio:
            itens_patrimonio = itens_patrimonio.filter(data_aquisicao__gte=data_aquisicao_inicio)
        if data_aquisicao_fim:
            itens_patrimonio = itens_patrimonio.filter(data_aquisicao__lte=data_aquisicao_fim)

    itens_patrimonio = itens_patrimonio.order_by('nome')

    context = {
        'itens_patrimonio': itens_patrimonio,
        'filter_form': filter_form,
        'active_page': 'patrimonio',
    }
    return render(request, 'patrimonio/listar_itens_patrimonio.html', context)


@login_required
def adicionar_item_patrimonio(request):
    if request.method == 'POST':
        form = ItemPatrimonioForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.usuario_registro = request.user 
            item.data_registro = timezone.now() 
            item.responsavel_atual = request.user 
            item.save()

            messages.success(request, f'Item de patrimônio "{item.nome}" adicionado com sucesso!')
            return redirect('patrimonio:detalhar_item_patrimonio', pk=item.pk)
        else:
            messages.error(request, 'Erro ao adicionar item de patrimônio. Verifique os campos.')
    else:
        form = ItemPatrimonioForm()

    context = {
        'form': form,
        'active_page': 'patrimonio_adicionar',
    }
    return render(request, 'patrimonio/adicionar_item_patrimonio.html', context)


@login_required
def editar_item_patrimonio(request, pk):
    item = get_object_or_404(ItemPatrimonio, pk=pk)
    if request.method == 'POST':
        form = ItemPatrimonioForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f'Item de patrimônio "{item.nome}" atualizado com sucesso!')
            return redirect('patrimonio:listar_itens_patrimonio')
        else:
            messages.error(request, 'Erro ao atualizar item de patrimônio. Verifique os campos.')
    else:
        form = ItemPatrimonioForm(instance=item)

    context = {
        'form': form,
        'item': item,
        'active_page': 'patrimonio',
    }
    return render(request, 'patrimonio/editar_item_patrimonio.html', context)


@login_required
def excluir_item_patrimonio(request, pk):
    item = get_object_or_404(ItemPatrimonio, pk=pk)
    if request.method == 'POST':
        item.delete()
        messages.success(request, f'Item de patrimônio "{item.nome}" excluído com sucesso.')
        return redirect('patrimonio:listar_itens_patrimonio')
    messages.error(request, 'A exclusão deve ser feita via POST.')
    return redirect('patrimonio:listar_itens_patrimonio')


@login_required
def baixar_item_patrimonio(request, pk):
    item = get_object_or_404(ItemPatrimonio, pk=pk)

    if item.status == 'BAIXADO':
        messages.warning(request, f'O item "{item.nome}" já está baixado.')
        return redirect('patrimonio:listar_itens_patrimonio')

    if request.method == 'POST':
        form = BaixaPatrimonioForm(request.POST)
        if form.is_valid():
            item.status = 'BAIXADO'
            item.data_baixa = form.cleaned_data['data_baixa']
            item.motivo_baixa = form.cleaned_data['motivo_baixa']
            item.save()

            MovimentacaoPatrimonio.objects.create(
                item=item,
                tipo_movimentacao='BAIXA',
                data_movimentacao=timezone.now(),
                localizacao_origem=item.localizacao,
                responsavel_origem=item.responsavel_atual,
                observacoes=f"Item baixado: {item.motivo_baixa}",
                usuario_registro=request.user
            )

            messages.success(request, f'Item "{item.nome}" baixado com sucesso.')
            return redirect('patrimonio:listar_itens_patrimonio')
        else:
            messages.error(request, 'Erro ao baixar o item. Verifique os campos.')
    else:
        form = BaixaPatrimonioForm(initial={'data_baixa': timezone.now().date()})

    context = {
        'form': form,
        'item': item,
        'active_page': 'patrimonio',
    }
    return render(request, 'patrimonio/baixar_item_patrimonio.html', context)


@login_required
def transferir_item_patrimonio(request, pk):
    item = get_object_or_404(ItemPatrimonio, pk=pk)

    if request.method == 'POST':
        form = TransferenciaPatrimonioForm(request.POST)
        if form.is_valid():
            nova_localizacao = form.cleaned_data.get('nova_localizacao')
            novo_responsavel = form.cleaned_data.get('novo_responsavel')
            motivo_transferencia = form.cleaned_data.get('motivo_transferencia')

            localizacao_anterior = item.localizacao
            responsavel_anterior = item.responsavel_atual

            if nova_localizacao:
                item.localizacao = nova_localizacao
            if novo_responsavel:
                item.responsavel_atual = novo_responsavel
            item.status = 'TRANSFERIDO'
            item.save()

            MovimentacaoPatrimonio.objects.create(
                item=item,
                tipo_movimentacao='TRANSFERENCIA',
                data_movimentacao=timezone.now(),
                localizacao_origem=localizacao_anterior,
                localizacao_destino=nova_localizacao,
                responsavel_origem=responsavel_anterior,
                responsavel_destino=novo_responsavel,
                observacoes=f"Item transferido. Motivo: {motivo_transferencia}",
                usuario_registro=request.user
            )

            messages.success(request, f'Item "{item.nome}" transferido com sucesso.')
            return redirect('patrimonio:listar_itens_patrimonio')
        else:
            messages.error(request, 'Erro ao transferir o item. Verifique os campos.')
    else:
        form = TransferenciaPatrimonioForm(initial={
            'localizacao_atual': item.localizacao,
            'responsavel_atual': item.responsavel_atual
        })

    context = {
        'form': form,
        'item': item,
        'active_page': 'patrimonio',
    }
    return render(request, 'patrimonio/transferir_item_patrimonio.html', context)


@login_required
def listar_movimentacoes(request):
    form = MovimentacaoFilterForm(request.GET)

    movimentacoes_filtradas = MovimentacaoPatrimonio.objects.all()

    if form.is_valid():
        q = form.cleaned_data.get('q')
        tipo_movimentacao = form.cleaned_data.get('tipo_movimentacao')
        localizacao_origem = form.cleaned_data.get('localizacao_origem')
        localizacao_destino = form.cleaned_data.get('localizacao_destino')
        data_inicio = form.cleaned_data.get('data_inicio')
        data_fim = form.cleaned_data.get('data_fim')

        if q:
            movimentacoes_filtradas = movimentacoes_filtradas.filter(
                Q(item__nome__icontains=q) | 
                Q(item__codigo_patrimonial__icontains=q) |
                Q(observacoes__icontains=q)
            )
        if tipo_movimentacao:
            movimentacoes_filtradas = movimentacoes_filtradas.filter(tipo_movimentacao=tipo_movimentacao)
        if localizacao_origem:
            movimentacoes_filtradas = movimentacoes_filtradas.filter(localizacao_origem=localizacao_origem)
        if localizacao_destino:
            movimentacoes_filtradas = movimentacoes_filtradas.filter(localizacao_destino=localizacao_destino)
        if data_inicio:
            movimentacoes_filtradas = movimentacoes_filtradas.filter(data_movimentacao__gte=data_inicio)
        if data_fim:
            movimentacoes_filtradas = movimentacoes_filtradas.filter(data_movimentacao__lt=data_fim + timedelta(days=1))

    movimentacoes_filtradas = movimentacoes_filtradas.order_by('-data_movimentacao')

    paginator = Paginator(movimentacoes_filtradas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'form': form,
        'movimentacoes': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages,
    }

    return render(request, 'patrimonio/listar_movimentacoes.html', context)


@login_required
# @permission_required('patrimonio.view_itempatrimonio', raise_exception=True) # Exemplo de permissão
def gerar_relatorio_patrimonio(request):
    # Inicializa o FilterSet com os dados da requisição GET e o queryset base
    item_filter = ItemPatrimonioFilter(request.GET, queryset=ItemPatrimonio.objects.all().order_by('codigo_patrimonial'))
    
    # O queryset filtrado é acessível via .qs
    itens_patrimonio = item_filter.qs 

    # O formulário que o crispy_forms precisa para renderizar os campos é acessível via .form
    filter_form = item_filter.form 
    
    # Os prints de depuração podem ser removidos, mas se quiser mantê-los para verificação:
    # if not filter_form.is_valid():
    #     print("DEBUG: Erros de validação do filter_form:", filter_form.errors)
    # else:
    #     print("DEBUG: filter_form é válido. Dados:", filter_form.cleaned_data)

    context = {
        'itens_patrimonio': itens_patrimonio,
        'filter_form': filter_form, # Agora filter_form é uma instância de django.forms.Form
        'title': 'Relatório de Patrimônio',
        'active_page': 'relatorio_patrimonio',
    }
    return render(request, 'patrimonio/relatorio_patrimonio.html', context)


@login_required
# @permission_required('patrimonio.view_itempatrimonio', raise_exception=True)
def exportar_patrimonio_csv(request):
    itens_patrimonio = ItemPatrimonio.objects.all().order_by('codigo_patrimonial')

    filter_form = PatrimonioFilterForm(request.GET)
    if filter_form.is_valid():
        search_query = filter_form.cleaned_data.get('search_query')
        categoria = filter_form.cleaned_data.get('categoria')
        localizacao = filter_form.cleaned_data.get('localizacao')
        estado_conservacao = filter_form.cleaned_data.get('estado_conservacao')
        status = filter_form.cleaned_data.get('status')
        data_aquisicao_inicio = filter_form.cleaned_data.get('data_aquisicao_inicio')
        data_aquisicao_fim = filter_form.cleaned_data.get('data_aquisicao_fim')

        if search_query:
            itens_patrimonio = itens_patrimonio.filter(
                Q(nome__icontains=search_query) |
                Q(codigo_patrimonial__icontains=search_query) |
                Q(numero_serie__icontains=search_query) |
                Q(descricao__icontains=search_query)
            )
        if categoria:
            itens_patrimonio = itens_patrimonio.filter(categoria=categoria)
        if localizacao:
            itens_patrimonio = itens_patrimonio.filter(localizacao=localizacao)
        if estado_conservacao:
            itens_patrimonio = itens_patrimonio.filter(estado_conservacao=estado_conservacao)
        if status:
            itens_patrimonio = itens_patrimonio.filter(status=status)
        if data_aquisicao_inicio:
            itens_patrimonio = itens_patrimonio.filter(data_aquisicao__gte=data_aquisicao_inicio)
        if data_aquisicao_fim:
            itens_patrimonio = itens_patrimonio.filter(data_aquisicao__lte=data_aquisicao_fim)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="relatorio_patrimonio_csv_{timezone.localdate()}.csv"'

    writer = csv.writer(response)

    headers = [
        'Código Patrimonial', 'Nome', 'Número de Série', 'Descrição',
        'Data de Aquisição', 'Valor de Aquisição (R$)', 'Estado de Conservação',
        'Status', 'Categoria', 'Localização', 'Responsável Atual',
        'Data da Baixa', 'Motivo da Baixa', 'Data de Registro', 'Última Atualização'
    ]
    writer.writerow(headers) # CORRIGIDO: Era sheet.append(headers)

    for item in itens_patrimonio:
        writer.writerow([
            item.codigo_patrimonial,
            item.nome,
            item.numero_serie,
            item.descricao,
            item.data_aquisicao.strftime('%Y-%m-%d') if item.data_aquisicao else '',
            f"{item.valor_aquisicao:.2f}".replace('.', ','),
            item.get_estado_conservacao_display(),
            item.get_status_display(),
            item.categoria.nome if item.categoria else '',
            item.localizacao.nome if item.localizacao else '',
            item.responsavel_atual.get_full_name() if item.responsavel_atual else item.responsavel_atual.username if item.responsavel_atual else '',
            item.data_baixa.strftime('%Y-%m-%d') if item.data_baixa else '',
            item.motivo_baixa if item.motivo_baixa else '',
            item.data_registro.strftime('%Y-%m-%d %H:%M:%S') if item.data_registro else '',
            item.data_ultima_atualizacao.strftime('%Y-%m-%d %H:%M:%S') if item.data_ultima_atualizacao else ''
        ])
    return response


@login_required
# @permission_required('patrimonio.view_itempatrimonio', raise_exception=True)
def exportar_patrimonio_excel(request):
    itens_patrimonio = ItemPatrimonio.objects.all().order_by('codigo_patrimonial')

    filter_form = PatrimonioFilterForm(request.GET)
    if filter_form.is_valid():
        search_query = filter_form.cleaned_data.get('search_query')
        categoria = filter_form.cleaned_data.get('categoria')
        localizacao = filter_form.cleaned_data.get('localizacao')
        estado_conservacao = filter_form.cleaned_data.get('estado_conservacao')
        status = filter_form.cleaned_data.get('status')
        data_aquisicao_inicio = filter_form.cleaned_data.get('data_aquisicao_inicio')
        data_aquisicao_fim = filter_form.cleaned_data.get('data_aquisicao_fim')

        if search_query:
            itens_patrimonio = itens_patrimonio.filter(
                Q(nome__icontains=search_query) |
                Q(codigo_patrimonial__icontains=search_query) |
                Q(numero_serie__icontains=search_query) |
                Q(descricao__icontains=search_query)
            )
        if categoria:
            itens_patrimonio = itens_patrimonio.filter(categoria=categoria)
        if localizacao:
            itens_patrimonio = itens_patrimonio.filter(localizacao=localizacao)
        if estado_conservacao:
            itens_patrimonio = itens_patrimonio.filter(estado_conservacao=estado_conservacao)
        if status:
            itens_patrimonio = itens_patrimonio.filter(status=status)
        if data_aquisicao_inicio:
            itens_patrimonio = itens_patrimonio.filter(data_aquisicao__gte=data_aquisicao_inicio)
        if data_aquisicao_fim:
            itens_patrimonio = itens_patrimonio.filter(data_aquisicao__lte=data_aquisicao_fim)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_patrimonio_excel_{timezone.localdate()}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de Patrimônio"

    headers = [
        'Código Patrimonial', 'Nome', 'Número de Série', 'Descrição',
        'Data de Aquisição', 'Valor de Aquisição (R$)', 'Estado de Conservação',
        'Status', 'Categoria', 'Localização', 'Responsável Atual',
        'Data da Baixa', 'Motivo da Baixa', 'Data de Registro', 'Última Atualização'
    ]
    sheet.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for item in itens_patrimonio:
        row_data = [
            item.codigo_patrimonial,
            item.nome,
            item.numero_serie,
            item.descricao,
            item.data_aquisicao.strftime('%Y-%m-%d') if item.data_aquisicao else '',
            float(item.valor_aquisicao) if item.valor_aquisicao else 0,
            item.get_estado_conservacao_display(),
            item.get_status_display(),
            item.categoria.nome if item.categoria else '',
            item.localizacao.nome if item.localizacao else '',
            item.responsavel_atual.get_full_name() if item.responsavel_atual else item.responsavel_atual.username if item.responsavel_atual else '',
            item.data_baixa.strftime('%Y-%m-%d') if item.data_baixa else '',
            item.motivo_baixa if item.motivo_baixa else '',
            item.data_registro.strftime('%Y-%m-%d %H:%M:%S') if item.data_registro else '',
            item.data_ultima_atualizacao.strftime('%Y-%m-%d %H:%M:%S') if item.data_ultima_atualizacao else ''
        ]
        sheet.append(row_data)

    for col_num, column_data in enumerate(sheet.columns, 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in column_data:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

        if headers[col_num - 1] == 'Valor de Aquisição (R$)':
            for cell in column_data[1:]:
                cell.number_format = '"R$"#,##0.00'

    workbook.save(response)
    return response


@login_required
def detalhar_item_patrimonio(request, pk):
    item = get_object_or_404(ItemPatrimonio, pk=pk)
    context = {
        'item': item,
        'title': f'Detalhes do Item: {item.nome}',
        'active_page': 'listar_itens_patrimonio',
    }
    return render(request, 'patrimonio/detalhar_item_patrimonio.html', context)

# ... (seus imports no topo do arquivo) ...
from django.db.models import Q # Para consultas OR
from django.utils import timezone # Para timezone.now() e timedelta
from datetime import timedelta # Importe timedelta para o filtro de data
# Certifique-se que ColetaInventario está importado no topo, sem try/except aqui.
from .models import ItemPatrimonio, CategoriaPatrimonio, LocalizacaoPatrimonio, MovimentacaoPatrimonio, ColetaInventario 
# ... (restante dos seus imports e views anteriores) ...


@login_required
def relatorio_inventario_conferencia(request):
    """
    Gera um relatório para conferência de inventário, listando:
    - Total de itens cadastrados e ATIVOS.
    - Itens que foram coletados (inventariados) recentemente.
    - Itens ATIVOS que NÃO foram coletados recentemente.
    """
    # Define o período de "recentemente coletado" (ex: últimos 30 dias)
    # Ajuste este valor conforme a frequência do seu inventário.
    dias_para_considerar_coletado = 30 
    data_limite_coleta = timezone.now() - timedelta(days=dias_para_considerar_coletado)

    # 1. Total de Itens Cadastrados (Ativos)
    total_itens_ativos = ItemPatrimonio.objects.filter(status='ATIVO').count()
    
    # 2. Itens Coletados Recentemente (dentro do período definido)
    # Consideramos "coletado" APENAS se houver uma movimentação de INVENTARIO
    # OU um registro no modelo ColetaInventario dentro do período.
    
    # IDs de itens coletados via Movimentacao (tipo INVENTARIO)
    ids_coletados_via_movimentacao = MovimentacaoPatrimonio.objects.filter(
        tipo_movimentacao='INVENTARIO',
        data_movimentacao__gte=data_limite_coleta
    ).values_list('item__id', flat=True)

    # IDs de itens coletados via modelo ColetaInventario
    ids_coletados_via_coleta_inventario = ColetaInventario.objects.filter(
        data_coleta__gte=data_limite_coleta
    ).values_list('item_patrimonio__id', flat=True)
    
    # Combina os IDs das fontes explícitas de inventário.
    # Removida a verificação de 'data_ultima_atualizacao' para evitar falsos positivos de "coleta".
    ids_todos_coletados_recentemente = list(set(
        list(ids_coletados_via_movimentacao) + 
        list(ids_coletados_via_coleta_inventario)
    ))

    # Pega os objetos ItemPatrimonio correspondentes aos IDs coletados recentemente
    itens_coletados_recentemente_objetos = ItemPatrimonio.objects.filter(
        id__in=ids_todos_coletados_recentemente,
        status='ATIVO' 
    ).order_by('codigo_patrimonial')
    
    total_coletados_recentemente = itens_coletados_recentemente_objetos.count()

    # 3. Itens Cadastrados (Ativos) e NÃO Coletados (no período)
    # Estes são os itens ATIVOS que NÃO estão na lista combinada de IDs coletados explicitamente.
    itens_nao_coletados = ItemPatrimonio.objects.filter(
        status='ATIVO'
    ).exclude(
        id__in=ids_todos_coletados_recentemente
    ).order_by('codigo_patrimonial')
    
    total_nao_coletados = itens_nao_coletados.count()

    context = {
        'total_itens_ativos': total_itens_ativos,
        
        'itens_coletados': itens_coletados_recentemente_objetos,
        'total_coletados_recentemente': total_coletados_recentemente,
        
        'itens_nao_coletados': itens_nao_coletados,
        'total_nao_coletados': total_nao_coletados,
        
        'dias_coleta': dias_para_considerar_coletado,
        'data_limite': data_limite_coleta.strftime('%d/%m/%Y %H:%M'),
        'active_page': 'relatorio_inventario_conferencia',
        'title': 'Relatório de Conferência de Inventário',
    }
    return render(request, 'patrimonio/relatorio_inventario_conferencia.html', context)

# - - --- Etiquetas Patrimonio -----

@login_required
def imprimir_etiquetas_view(request):
    item_ids_str = request.GET.get('ids', '')
    if not item_ids_str:
        messages.error(request, "Nenhum item selecionado para impressão.")
        # Redireciona de volta para a página de relatório se nenhum ID for fornecido
        return redirect('relatorio_patrimonio') 

    # Converte a string de IDs para uma lista de inteiros
    item_ids = [int(item_id) for item_id in item_ids_str.split(',') if item_id.isdigit()]
    
    # Busca os itens de patrimônio selecionados
    itens_para_imprimir = ItemPatrimonio.objects.filter(id__in=item_ids)

    etiquetas_data = []
    for item in itens_para_imprimir:
        qr_image_base64 = None
        
        # O dado que será codificado no QR Code. 
        # Sugestão: a URL completa para a página de detalhes do item no seu sistema.
        # Substitua 'patrimonio_detalhe' pelo nome da URL para a página de detalhes de um único item.
        # Se você não tem uma URL de detalhe, pode usar o código patrimonial: f"Código: {item.codigo_patrimonial}"
        # Exemplo: Se você tem uma URL como path('patrimonio/<int:item_id>/', views.detalhe_patrimonio, name='patrimonio_detalhe'), use 'patrimonio_detalhe'
        try:
            # Tenta gerar a URL para o detalhe do item. Adapte 'patrimonio_detalhe' ao seu urls.py
            qr_data = request.build_absolute_uri(reverse('detalhe_patrimonio', args=[item.id])) 
        except Exception:
            # Se a URL não puder ser gerada (ex: URL 'detalhe_patrimonio' não existe), use o código patrimonial
            qr_data = f"Código: {item.codigo_patrimonial}"

        # Geração do QR Code
        qr = qrcode.QRCode(
            version=1, # Tamanho do QR Code (1 a 40)
            error_correction=qrcode.constants.ERROR_CORRECT_L, # Nível de correção de erro (L, M, Q, H)
            box_size=10, # Tamanho de cada "caixa" do QR Code
            border=4, # Largura da borda branca
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        # Salva a imagem do QR Code em memória e codifica para Base64
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        qr_image_base64 = base64.b64encode(buffer.getvalue()).decode()
        
        etiquetas_data.append({
            'nome': item.nome,
            'codigo_patrimonial': item.codigo_patrimonial,
            'qr_code_base64': qr_image_base64,
        })
    
    context = {
        'etiquetas_data': etiquetas_data,
    }
    return render(request, 'patrimonio/imprimir_etiquetas.html', context)

# - - --- Fim Etiquetas Patrimonio -----



@login_required
def exportar_inventario_conferencia_csv(request):
    """
    Exporta o Relatório de Conferência de Inventário para um arquivo CSV.
    Inclui itens coletados e itens não coletados.
    """
    # Define o período de "recentemente coletado" (o mesmo usado na view do relatório)
    dias_para_considerar_coletado = 30 
    data_limite_coleta = timezone.now() - timedelta(days=dias_para_considerar_coletado)

    # 1. Obter IDs de Itens Coletados Recentemente
    ids_coletados_via_movimentacao = MovimentacaoPatrimonio.objects.filter(
        tipo_movimentacao='INVENTARIO',
        data_movimentacao__gte=data_limite_coleta
    ).values_list('item__id', flat=True)

    ids_coletados_via_coleta_inventario = ColetaInventario.objects.filter(
        data_coleta__gte=data_limite_coleta
    ).values_list('item_patrimonio__id', flat=True)
    
    ids_todos_coletados_recentemente = list(set(
        list(ids_coletados_via_movimentacao) + 
        list(ids_coletados_via_coleta_inventario)
    ))

    # 2. Obter Itens Coletados e Não Coletados (ativos)
    itens_coletados = ItemPatrimonio.objects.filter(
        id__in=ids_todos_coletados_recentemente,
        status='ATIVO' 
    ).order_by('codigo_patrimonial')

    itens_nao_coletados = ItemPatrimonio.objects.filter(
        status='ATIVO'
    ).exclude(
        id__in=ids_todos_coletados_recentemente
    ).order_by('codigo_patrimonial')

    # Configura a resposta HTTP para CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="relatorio_inventario_conferencia_{timezone.localdate()}.csv"'

    writer = csv.writer(response, delimiter=';') # Use ponto e vírgula como delimitador para melhor compatibilidade com Excel

    # Cabeçalho para Itens Coletados
    writer.writerow(['RELATÓRIO DE CONFERÊNCIA DE INVENTÁRIO'])
    writer.writerow([f'Período de Coleta Recente: Últimos {dias_para_considerar_coletado} dias (Até {data_limite_coleta.strftime("%d/%m/%Y %H:%M")})'])
    writer.writerow([]) # Linha em branco
    writer.writerow(['ITENS COLETADOS RECENTEMENTE'])
    writer.writerow([
        'Código Patrimonial', 'Nome', 'Número de Série', 'Categoria', 'Localização Atual',
        'Status', 'Data Última Atualização'
    ])

    for item in itens_coletados:
        writer.writerow([
            item.codigo_patrimonial,
            item.nome,
            item.numero_serie,
            item.categoria.nome if item.categoria else '',
            item.localizacao.nome if item.localizacao else '',
            item.get_status_display(),
            item.data_ultima_atualizacao.strftime('%Y-%m-%d %H:%M:%S') if item.data_ultima_atualizacao else ''
        ])

    writer.writerow([]) # Linha em branco
    writer.writerow([]) # Linha em branco

    # Cabeçalho para Itens NÃO Coletados
    writer.writerow(['ITENS ATIVOS CADASTRADOS E NÃO COLETADOS NO PERÍODO'])
    writer.writerow([
        'Código Patrimonial', 'Nome', 'Número de Série', 'Categoria', 'Localização Atual',
        'Status', 'Data de Registro', 'Data Última Atualização'
    ])

    for item in itens_nao_coletados:
        writer.writerow([
            item.codigo_patrimonial,
            item.nome,
            item.numero_serie,
            item.categoria.nome if item.categoria else '',
            item.localizacao.nome if item.localizacao else '',
            item.get_status_display(),
            item.data_registro.strftime('%Y-%m-%d %H:%M:%S') if item.data_registro else '',
            item.data_ultima_atualizacao.strftime('%Y-%m-%d %H:%M:%S') if item.data_ultima_atualizacao else ''
        ])

    return response
